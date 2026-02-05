#!/usr/bin/env python3
"""
Example of reading and writing Excel files using openpyxl
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import argparse
import logging

# Files
root_files = './files/'

# Log files
log_file = root_files + 'log.txt'

# Debug
dry_run = False

def setup_debug():
    """Setup logging for debug mode""" 
    logger = logging.getLogger("")
    logging.basicConfig(level=logging.DEBUG)
    handler = logging.FileHandler(log_file)
    formatter = logging.Formatter("%(asctime)s %(levelname)s : %(message)s")
    handler.setFormatter(formatter)
    logger.addHandler(handler)


def write_excel_example(args):
    """Create and write data to an Excel file"""
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Add headers
    headers = ["Name", "Age", "City", "Date"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add data rows
    data = [
        ["Alice", 30, "New York", datetime.now()],
        ["Bob", 25, "San Francisco", datetime.now()],
        ["Charlie", 35, "Boston", datetime.now()],
    ]
    
    for row in data:
        ws.append(row)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    
    # Save the workbook
    wb.save('files/example.xlsx')
    print("✓ Excel file 'example.xlsx' created successfully")


def read_excel_example(args):
    """Read data from an Excel file"""
    try:
        # Load the workbook
        wb = load_workbook('files/example.xlsx')
        ws = wb.active
        
        print("\n✓ Reading from 'example.xlsx':")
        print("-" * 60)
        
        # Read and display all rows
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            print(row)
        
        print("-" * 60)
        
        # Access specific cells
        print(f"\nFirst person's name: {ws['A2'].value}")
        print(f"Number of data rows: {ws.max_row - 1}")
        
    except FileNotFoundError:
        print("Error: 'files/example.xlsx' not found. Please run write_excel_example() first.")


def update_excel_example(args):
    """Update existing Excel file"""
    try:
        # Load existing workbook
        wb = load_workbook('files/example.xlsx')
        ws = wb.active
        
        # Add a new row
        ws.append(["David", 28, "Chicago", datetime.now()])
        
        # Modify a cell
        ws['C2'] = "Los Angeles"  # Change Bob's city
        
        # Save changes
        wb.save('files/example.xlsx')
        print("\n✓ Excel file updated successfully")
        
    except FileNotFoundError:
        print("Error: 'files/example.xlsx' not found.")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(prog='rwxl',
                                        description='These functions will allow read and write excel files',
                                        epilog='See "%(prog)s help COMMAND" for help on a specific command.')
    parser.add_argument('--debug', '-d', action='count', help='Print debug output')
    parser.add_argument('--dry-run', '-dr', action='count', help='Execute a dry run')
    subparsers = parser.add_subparsers(dest='command', help='Available commands')

    write_parser = subparsers.add_parser('write', help='Write an Excel file')
    write_parser.set_defaults(func=write_excel_example)
    read_parser = subparsers.add_parser('read', help='Read an Excel file')
    read_parser.set_defaults(func=read_excel_example)

    args = parser.parse_args()

    if args.debug:
        setup_debug()

    if args.dry_run:
        dry_run = True

    if args.command:
        args.func(args)
    else:
        parser.print_help()
